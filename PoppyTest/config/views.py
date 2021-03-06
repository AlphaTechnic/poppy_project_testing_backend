import requests
import openpyxl
import numpy as np
from json import dumps
from haversine import haversine
from django.shortcuts import HttpResponse
from . import samples
import random


def price_to_int(price):
    # '30,000' -> 30000
    return int(price.replace(",", ""))


def distance_to_int(distance):
    # '300m' -> 300, '20km' -> 20000
    if distance[-2:] == 'km':
        return float(distance.replace("km", ""))*1000
    else:
        return float(distance.replace("m", ""))


def get_lat_lng(address):
    result = ""

    url = 'https://dapi.kakao.com/v2/local/search/address.json?query=' + address
    rest_api_key = '7d66b3ade18db0422c2f27baada16e45'
    header = {'Authorization': 'KakaoAK ' + rest_api_key}

    r = requests.get(url, headers=header)

    if r.status_code == 200:
        result_address = r.json()["documents"][0]["address"]
        result = float(result_address["x"]), float(result_address["y"])
    else:
        result = "ERROR[" + str(r.status_code) + "]"

    return result


def make_non_expert_address():
    expert_or_not = 0
    coordinate_start = 126.751286, 37.419093
    coordinate_end = 127.196322, 37.707367

    xvals = np.arange(coordinate_start[0], coordinate_end[0], step=0.009)
    yvals = np.arange(coordinate_start[1], coordinate_end[1], step=0.009)

    write_wb = openpyxl.Workbook()
    write_ws = write_wb.active

    for x in xvals:
        for y in yvals:
            result_address = ""

            url = 'https://dapi.kakao.com/v2/local/geo/coord2regioncode.json?x={}&y={}'.format(x, y)
            rest_api_key = '7d66b3ade18db0422c2f27baada16e45'
            header = {'Authorization': 'KakaoAK ' + rest_api_key}

            r = requests.get(url, headers=header)

            if r.status_code == 200:
                result_address = r.json()["documents"][0]["address_name"]
            else:
                result_address = "ERROR[" + str(r.status_code) + "]"

            if result_address[0:2] != '서울':
                continue

            ## 엑셀에 기록
            write_ws.append([x, y, result_address, expert_or_not])

    write_wb.save('non_expert_address.xlsx')

    print("비전문가들 주소 생성 완료")


def make_expert_address():
    expert_or_not = 1
    coordinate_start = 126.751286, 37.419093
    coordinate_end = 127.196322, 37.707367

    xvals = np.arange(coordinate_start[0], coordinate_end[0], step=0.08)
    yvals = np.arange(coordinate_start[1], coordinate_end[1], step=0.08)

    write_wb = openpyxl.Workbook()
    write_ws = write_wb.active

    for x in xvals:
        for y in yvals:
            result_address = ""

            url = 'https://dapi.kakao.com/v2/local/geo/coord2regioncode.json?x={}&y={}'.format(x, y)
            rest_api_key = '7d66b3ade18db0422c2f27baada16e45'
            header = {'Authorization': 'KakaoAK ' + rest_api_key}

            r = requests.get(url, headers=header)

            if r.status_code == 200:
                result_address = r.json()["documents"][0]["address_name"]
            else:
                result_address = "ERROR[" + str(r.status_code) + "]"

            if result_address[0:2] != '서울':
                continue

            ## 엑셀에 기록
            write_ws.append([x, y, result_address, expert_or_not])
    write_wb.save('expert_address.xlsx')
    print("전문가들 주소 생성 완료")


def get_distance(coordinate1, coordinate2):
    distance = haversine(coordinate1, coordinate2, unit='km')

    return distance


def get_experts_nearby(request):
    order_by = request.GET.get('order_by')

    address = request.GET.get('address')
    coordinate1 = get_lat_lng(address)

    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = openpyxl.load_workbook("expert_address.xlsx", data_only=True)
    load_ws = load_wb['Sheet']

    expert_list = []
    for row in load_ws.rows:
        x2 = row[0].value
        y2 = row[1].value
        address = row[2].value
        expert_or_not = row[3].value

        coordinate2 = x2, y2  # 펫시터 사는 곳 좌표
        new_querySet = x2, y2, address, get_distance(coordinate1, coordinate2), expert_or_not

        expert_list.append(new_querySet)

    nearest5_experts = sorted(expert_list, key=lambda x: x[3])[:5]

    info = dict()
    petsitters = []

    # 첫번째 펫시터와의 거리로 생성한 random number
    s = round(nearest5_experts[0][0])
    random.seed(s)
    ran_arr = [0, 1, 2, 3, 4]
    ran_arr = random.sample(ran_arr, 5)
    i = 0

    for i, petsitter in enumerate(nearest5_experts):
        petsitter_info = dict()
        # petsitter[4] : 전문가 여부
        petsitter_info["expert_or_not"] = 1

        # petsitter[2] : 주소
        petsitter_info["address"] = petsitter[2]

        # petsitter[3] : 떨어진 거리
        if petsitter[3] < 1:
            petsitter_info["distance"] = str(round(petsitter[3] * 1000)) + 'm'
        elif 1 <= petsitter[3] < 10:
            petsitter_info["distance"] = str(round(petsitter[3], 1)) + 'km'
        else:
            petsitter_info["distance"] = str(round(petsitter[3])) + 'km'

        # petsitter[5] : 5가지 유형 중 어떤 petsitter인가
        type = ran_arr[i]
        i += 1
        petsitter_info["type"] = type

        # title
        petsitter_info["title"] = samples.expert[type]["title"]
        # 방 사진
        petsitter_info["room_img"] = samples.expert[type]["room_img"]
        # 평점
        petsitter_info["score"] = samples.expert[type]["score"]
        # 가격
        petsitter_info["small_dog_cost"] = samples.expert[type]["small_dog_cost"]
        petsitters.append(petsitter_info)

    if order_by == 'price':
        sorted_petsitters = sorted(petsitters, key=lambda x: (x["small_dog_cost"][0], distance_to_int(x["distance"])))
        info["experts"] = sorted_petsitters
        return HttpResponse(dumps(info, ensure_ascii=False), content_type=u"application/json; charset=utf-8")

    info["experts"] = petsitters
    return HttpResponse(dumps(info, ensure_ascii=False), content_type=u"application/json; charset=utf-8")


def get_non_experts_nearby(request):
    order_by = request.GET.get('order_by')
    address = request.GET.get('address')
    coordinate1 = get_lat_lng(address)

    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = openpyxl.load_workbook("non_expert_address.xlsx", data_only=True)
    load_ws = load_wb['Sheet']

    non_expert_list = []
    for row in load_ws.rows:
        x2 = row[0].value
        y2 = row[1].value
        address = row[2].value
        expert_or_not = row[3].value

        coordinate2 = x2, y2  # 펫시터 사는 곳 좌표
        new_querySet = x2, y2, address, get_distance(coordinate1, coordinate2), expert_or_not

        non_expert_list.append(new_querySet)

    nearest5_non_experts = sorted(non_expert_list, key=lambda x: x[3])[:5]

    info = dict()
    petsitters = []

    # 첫번째 펫시터의 경도로 생성한 random number
    s = round(nearest5_non_experts[0][0]*10000)
    random.seed(s)
    ran_arr = [0, 1, 2, 3, 4]
    ran_arr = random.sample(ran_arr, 5)
    i = 0

    for i, petsitter in enumerate(nearest5_non_experts):
        petsitter_info = dict()
        # petsitter[4] : 전문가 여부
        petsitter_info["expert_or_not"] = 0

        # petsitter[2] : 주소
        petsitter_info["address"] = petsitter[2]

        # petsitter[3] : 떨어진 거리
        if petsitter[3] < 1:
            petsitter_info["distance"] = str(round(petsitter[3] * 1000)) + 'm'
        elif 1 <= petsitter[3] < 10:
            petsitter_info["distance"] = str(round(petsitter[3], 1)) + 'km'
        else:
            petsitter_info["distance"] = str(round(petsitter[3])) + 'km'

        # petsitter[5] : 5가지 유형 중 어떤 petsitter인가
        type = ran_arr[i]
        i += 1
        petsitter_info["type"] = type

        # title
        petsitter_info["title"] = samples.non_expert[type]["title"]
        # 방 사진
        petsitter_info["room_img"] = samples.non_expert[type]["room_img"]
        # 평점
        petsitter_info["score"] = samples.non_expert[type]["score"]
        # 가격
        petsitter_info["small_dog_cost"] = samples.non_expert[type]["small_dog_cost"]
        petsitters.append(petsitter_info)

    if order_by == 'price':
        sorted_petsitters = sorted(petsitters, key=lambda x: (x["small_dog_cost"][0], distance_to_int(x["distance"])))
        info["non_experts"] = sorted_petsitters
        return HttpResponse(dumps(info, ensure_ascii=False), content_type=u"application/json; charset=utf-8")

    info["non_experts"] = petsitters
    return HttpResponse(dumps(info, ensure_ascii=False), content_type=u"application/json; charset=utf-8")


def get_particular_expert(request, type):
    petsitter_info = dict()

    # petsitter[4] : 전문가 여부
    petsitter_info["expert_or_not"] = 1

    # 방 사진
    petsitter_info["room_img"] = samples.expert[type]["room_img"]
    # 이름
    petsitter_info["name"] = samples.expert[type]["name"]
    # 프로필 사진
    petsitter_info["profile_img"] = samples.expert[type]["profile_img"]
    # 강아지 가격
    petsitter_info["small_dog_cost"] = list(map(price_to_int, samples.expert[type]["small_dog_cost"]))
    petsitter_info["middle_dog_cost"] = list(map(price_to_int, samples.expert[type]["middle_dog_cost"]))
    petsitter_info["large_dog_cost"] = list(map(price_to_int, samples.expert[type]["large_dog_cost"]))
    # title
    petsitter_info["title"] = samples.expert[type]["title"]
    # content
    petsitter_info["content"] = samples.expert[type]["content"]
    # 강아지 정보
    petsitter_info["puppy"] = samples.expert[type]["puppy"]
    # 댓글
    petsitter_info["comment"] = samples.expert[type]["comment"]
    # 자격증
    petsitter_info["certification"] = samples.expert[type]["certification"]
    # 평점
    petsitter_info["score"] = samples.expert[type]["score"]

    return HttpResponse(dumps(petsitter_info, ensure_ascii=False), content_type=u"application/json; charset=utf-8")


def get_particular_non_expert(request, type):
    petsitter_info = dict()

    # petsitter[4] : 전문가 여부
    petsitter_info["expert_or_not"] = 0

    # 방 사진
    petsitter_info["room_img"] = samples.non_expert[type]["room_img"]
    # 이름
    petsitter_info["name"] = samples.non_expert[type]["name"]
    # 프로필 사진
    petsitter_info["profile_img"] = samples.non_expert[type]["profile_img"]
    # 강아지 가격
    petsitter_info["small_dog_cost"] = list(map(price_to_int, samples.non_expert[type]["small_dog_cost"]))
    petsitter_info["middle_dog_cost"] = list(map(price_to_int, samples.non_expert[type]["middle_dog_cost"]))
    petsitter_info["large_dog_cost"] = list(map(price_to_int, samples.non_expert[type]["large_dog_cost"]))
    # title
    petsitter_info["title"] = samples.non_expert[type]["title"]
    # content
    petsitter_info["content"] = samples.non_expert[type]["content"]
    # 강아지 정보
    petsitter_info["puppy"] = samples.non_expert[type]["puppy"]
    # 댓글
    petsitter_info["comment"] = samples.non_expert[type]["comment"]
    # 평점
    petsitter_info["score"] = samples.non_expert[type]["score"]

    return HttpResponse(dumps(petsitter_info, ensure_ascii=False), content_type=u"application/json; charset=utf-8")